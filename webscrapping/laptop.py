from bs4 import BeautifulSoup
from selenium import webdriver
import time
import openpyxl as op
# FILE NAME GOES HERE
file = 'data.xlsx'
_name = 1
_price = 2
_rating = 3
data = {}
data['name'] = []
data['price'] = []
data['ratings'] = []
#CHROME WEBDRIVER IS REQUIRED
browser = webdriver.Chrome('D:\\Work\\Work2\\chromedriver.exe')
browser.get('https://www.amazon.in/s?k=laptop&ref=nb_sb_noss')
pages = int(input('Enter no of pages : '))
print('Processing...... ')
for i in range(pages):
    time.sleep(2)
    page = browser.page_source
    soup = BeautifulSoup(page,features = 'lxml')
    time.sleep(5)
    for j in soup.findAll(name='div', attrs={'class': 's-include-content-margin s-border-bottom s-latency-cf-section'}):
        a = j.find(name='span', attrs={
                   'class': 'a-size-medium a-color-base a-text-normal'})
        p = j.find(name='span', attrs={'class': 'a-offscreen'})
        r = j.find(name='span', attrs={'class': 'a-icon-alt'})
        data['name'].append(a.text)
        if p != None:
            data['price'].append(p.text)
        else:
            data['price'].append('None')
        if r != None:
            rat = r.text.split(' ')[0]
            data['ratings'].append(rat)
        else:
            data['ratings'].append('None')
    time.sleep(2)
    print(f'Page {i+1} done.')
    time.sleep(2)
    browser.find_element_by_xpath(f'//a[text()="Next"]').click()
workbook = op.load_workbook(file)
sheet = workbook['Sheet1']
print('Total number of laptop data : ',len(data['name']))
for i in range(len(data['name'])):
    row = 2+i
    if sheet.cell(row, _name).value == None:
        sheet.cell(row, _name).value = data['name'][i]
        sheet.cell(row, _price).value = data['price'][i]
        sheet.cell(row, _rating).value = data['ratings'][i]
    else:
        print('Please check your file.')
workbook.save(file)
workbook.close()
print('File saved Successfully!')