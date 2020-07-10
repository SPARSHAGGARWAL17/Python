import random
import string
import openpyxl as op
import getpass
# file name goes here
file = 'data.xlsx'
_acc = 1
_name = 2
_pas = 3
_amount = 4
def workbooks():
    workbook = op.load_workbook(file)
    sheet = workbook['Sheet1']
    row = sheet.max_row+1
    return(workbook,sheet,row)
def create_account():
    workbook,sheet,row = workbooks()
    account_list = []
    for i in range(1,row):
        cell = sheet.cell(i,1)
        account_list.append(cell)
    print('~ Create Account in Bank. ~')
    name = input('Enter your name : ')
    letters = string.ascii_letters
    numbers = string.digits
    while True:
        account_no = random.randint(100000000,999999999)
        if account_no in account_list:
            continue
        else:
            sheet.cell(row,_acc).value = account_no
            break
    password = ''.join(random.choice(letters+numbers) for i in range(15))
    sheet.cell(row,_pas).value = password
    sheet.cell(row,_name).value = name
    sheet.cell(row,_amount).value = 0
    print('\nHello, ',name.upper())
    print('Use the following details to Login.')
    print('-----------------------------------')
    print('Account Number : ',account_no)
    print('Password : ',password)
    print('-----------------------------------')
    workbook.save(file)
def login():
    workbook,sheet,row = workbooks()
    print('~ Enter details to login ~')
    while True:
        login = (False,)
        account_no = input('Enter your Account number : ')
        password = getpass.getpass('Enter password : ','*')
        for i in range(1,row):
            acc = sheet.cell(i,_acc).value
            pas = sheet.cell(i,_pas).value
            if account_no == str(acc):
                if password == str(pas):
                    login = (True,i)
                    break
        if login[0]:
            print('Login Successful!')
            print('-----------------------------------')
            name = sheet.cell(login[1],_name).value
            print('Hello, ',name.upper())
            while True:
                choice = input('Enter your choice: \n 1. Transactions \n 2. Transfer \n 3. Change Password \n 4. Delete Account \n 5. Logout : ')
                if choice == '1':
                    transact(login)
                    continue
                elif choice =='2':
                    transfer(login)
                    continue
                elif choice == '3':
                    password = input("Enter new Password : ")
                    sheet.cell(login[1],_pas).value = password
                    print('Password Change Successfully!')
                    print('-----------------------------')
                    workbook.save(file)
                elif choice =='4':
                    delete(login)
                    print('Account Deleted Successfully!')
                    break
                elif choice == '5':
                    break
                else:
                    print('Invalid Choice')
                    continue
        else:
            print('Invalid User')
        break
def transact(login):
    workbook,sheet,row = workbooks()
    amount = sheet.cell(login[1],_amount).value
    print('-----------------------------------')
    print("Amount Available : ",amount)
    while True:
        choice = input('\nSelect Choice: \n 1. Deposit \n 2. Withdraw : ')
        if choice =='1':
            while True:
                try:
                    deposit = int(input('Enter amount to Deposit or enter (0) to go back : '))
                    sheet.cell(login[1],_amount).value = amount+deposit
                    print('Deposited Successfully.')
                    print('-----------------------------------')
                    print("Amount Available : ",amount+deposit)
                    workbook.save(file)
                    break
                except:
                    print('Invalid Value!')
                    continue
            break
        elif choice == '2':
            while True:
                try:
                    withdraw = int(input("Enter amount to Withdraw or enter (0) to go back : ")) 
                    if 0 <= withdraw < amount :
                        sheet.cell(login[1],_amount).value = amount-withdraw
                        print('Withdraw Successful!')
                        print('-----------------------------------')
                        print('Available Amount : ',amount-withdraw)
                        workbook.save(file)
                        break
                    else:
                        print('Invalid Amount!')
                        break
                except:
                    print('Invalid Amount!')
            break
def transfer(login):
    workbook,sheet,row = workbooks()
    print('-----------------------------------')
    while True:
        amount = sheet.cell(login[1],_amount).value
        print('Available Amount : ',amount)
        user = (False,0)
        try:
            account_no = int(input('Enter account number to Transfer : '))
            for i in range(1,row):
                acc = sheet.cell(i,_acc).value
                name = sheet.cell(i,_name).value
                if str(account_no) == str(acc):
                    print("BANK USER FOUND!")
                    print('Bank User : ',name.upper())
                    user = (True,i)
            if user[0]:
                while True:
                    amount_user = int(input('Enter amount to Transfer: '))
                    if amount_user >= amount:
                        print('Invalid Amount')
                        continue
                    else:
                        break
                sheet.cell(login[1],_amount).value = amount-amount_user
                sheet.cell(user[1],_amount).value+=amount_user
                print('Transfer Successful!')
                print('-----------------------------------')
                print('Available Amount : ',amount-amount_user)
                workbook.save(file)
                break
            else:
                print('Invalid User!')
                continue
        except:
            print('Invalid acc.')
def delete(login):
    workbook,sheet,row = workbooks()
    while True:
        choice = input('Are you sure to delete your account? [Y/N] ').lower()
        if choice == 'y':
            sheet.cell(login[1],_acc).value = None
            sheet.cell(login[1],_pas).value = None
            sheet.cell(login[1],_name).value = None
            sheet.cell(login[1],_amount).value = None
            workbook.save(file)
            break
        elif choice == 'n':
            break
        else:
            continue
print('===========================')
print('  ~ WELCOME TO THE BANK ~  ')
print('===========================')
while True:
    try:
        ch = int(input("Enter your choice: \n 1. Create Account \n 2. Login \n 3. Exit : "))
        if ch ==1:
            create_account()
            continue
        elif ch == 2:
            login()
            continue
        elif ch ==3:
            break
        else:
            print('Invalid Choice!')
            continue
    except:
        print('Invalid Choice!')
        continue